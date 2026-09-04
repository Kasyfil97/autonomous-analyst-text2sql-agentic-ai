"""
check_tables_in_db.py

Reads table names from adhoc_table.xls and checks whether each table
exists in schema_tables (the KB catalog on PG_HOST/postgres).

Usage:
    python check_tables_in_db.py [--input FILE] [--output FILE]
"""

import argparse
import os
import sys

import psycopg2
import xlrd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()


def get_conn():
    return psycopg2.connect(
        host=os.getenv("PG_HOST", "localhost"),
        port=int(os.getenv("PG_PORT", 5432)),
        dbname=os.getenv("PG_DBNAME", "postgres"),
        user=os.getenv("PG_USER"),
        password=os.getenv("PG_PASSWORD"),
    )


def load_xls_rows(xls_path: str) -> list[tuple[str, int, int]]:
    """Return (table_name, issues_created, issues_resolved) for each data row."""
    wb = xlrd.open_workbook(xls_path)
    sh = wb.sheet_by_index(0)
    rows = []
    for i in range(2, sh.nrows):  # rows 0-1 are headers/totals
        name = str(sh.cell_value(i, 0)).strip()
        if not name or name.lower() == "(none)":
            continue
        created = int(sh.cell_value(i, 1)) if sh.ncols > 1 and sh.cell_value(i, 1) else 0
        resolved = int(sh.cell_value(i, 2)) if sh.ncols > 2 and sh.cell_value(i, 2) else 0
        rows.append((name, created, resolved))
    return rows


def fetch_kb_tables(conn) -> dict[str, tuple[str, str]]:
    """Return {lower_table_name: (table_name, source_schema)} from schema_tables KB."""
    cur = conn.cursor()
    cur.execute("SELECT table_name, source_schema FROM schema_tables")
    result = {}
    for table_name, source_schema in cur.fetchall():
        result[table_name.lower()] = (table_name, source_schema or "")
    return result


def write_xlsx(
    output_path: str,
    rows: list[tuple[str, int, int]],
    kb: dict[str, tuple[str, str]],
) -> tuple[int, int]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Table Checklist"

    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    headers = ["Table Name", "Issues Created", "Issues Resolved", "In KB?", "KB Source Schema"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(1, col).font = bold
        ws.cell(1, col).alignment = center

    found = not_found = 0
    for name, created, resolved in rows:
        match = kb.get(name.lower())
        in_kb = "YES" if match else "NO"
        kb_schema = match[1] if match else ""
        fill = green if match else red
        ws.append([name, created, resolved, in_kb, kb_schema])
        r = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(r, col).fill = fill
        if match:
            found += 1
        else:
            not_found += 1

    for col in range(1, len(headers) + 1):
        max_len = max(len(str(ws.cell(row, col).value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 60)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    ws2.cell(1, 1).font = bold
    ws2.cell(1, 2).font = bold
    ws2.append(["Total tables checked", found + not_found])
    ws2.append(["Found in KB (YES)", found])
    ws2.append(["Not in KB (NO)", not_found])
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 20

    wb.save(output_path)
    return found, not_found


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="adhoc_table.xls")
    parser.add_argument("--output", default="adhoc_table_checklist.xlsx")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: {args.input} not found")
        sys.exit(1)

    print(f"Reading: {args.input}")
    rows = load_xls_rows(args.input)
    print(f"  {len(rows)} table names loaded")

    host = os.getenv("PG_HOST", "localhost")
    port = os.getenv("PG_PORT", 5432)
    dbname = os.getenv("PG_DBNAME", "postgres")
    print(f"Connecting to KB: {host}:{port}/{dbname}")
    conn = get_conn()

    print("Fetching schema_tables from KB...")
    kb = fetch_kb_tables(conn)
    conn.close()
    print(f"  {len(kb)} tables in KB")

    print(f"Writing: {args.output}")
    found, not_found = write_xlsx(args.output, rows, kb)

    print(f"\nDone.")
    print(f"  YES (in KB):  {found}")
    print(f"  NO  (missing): {not_found}")
    print(f"  Output: {args.output}")


if __name__ == "__main__":
    main()
