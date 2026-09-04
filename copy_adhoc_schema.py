"""
copy_adhoc_schema.py

Reads the YES table names from adhoc_table_checklist.xlsx, then copies all
schema_* tables from the public schema into a new target schema — filtering
schema_tables and schema_columns to only include those YES tables.

Usage:
    python copy_adhoc_schema.py [--schema TARGET_SCHEMA] [--checklist FILE] [--drop]

Options:
    --schema    Name of the new schema to create (default: adhoc)
    --checklist Path to the checklist xlsx (default: adhoc_table_checklist.xlsx)
    --drop      Drop and recreate the target schema if it already exists
"""

import argparse
import os
import sys

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

# schema_* tables that have a table_name column → filter by YES list
FILTERED_TABLES = ["schema_tables", "schema_columns"]

# schema_* tables that are indexes/metadata → copy entirely
FULL_COPY_TABLES = [
    "schema_tables_bm25",
    "schema_tables_bm25_meta",
    "schema_columns_bm25",
    "schema_columns_bm25_meta",
]


def get_conn():
    return psycopg2.connect(
        host=os.getenv("PG_HOST", "localhost"),
        port=int(os.getenv("PG_PORT", 5432)),
        dbname=os.getenv("PG_DBNAME", "postgres"),
        user=os.getenv("PG_USER"),
        password=os.getenv("PG_PASSWORD"),
    )


def read_yes_tables(checklist_path: str) -> set[str]:
    """Return lowercase table names where 'In KB?' == 'YES'."""
    wb = load_workbook(checklist_path, read_only=True)
    ws = wb["Table Checklist"]
    yes = set()
    header = None
    inkb_col = None
    name_col = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = list(row)
            inkb_col = header.index("In KB?")
            name_col = header.index("Table Name")
            continue
        if row[inkb_col] == "YES":
            yes.add(str(row[name_col]).lower())
    wb.close()
    return yes


def table_columns(cur, table_name: str) -> list[str]:
    cur.execute(
        """
        SELECT column_name FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s
        ORDER BY ordinal_position
        """,
        (table_name,),
    )
    return [r[0] for r in cur.fetchall()]


def copy_filtered(cur, src_table: str, tgt_schema: str, yes_names: set[str]) -> int:
    """Copy rows from public.src_table where table_name is in yes_names."""
    cols = table_columns(cur, src_table)
    col_list = ", ".join(f'"{c}"' for c in cols)

    cur.execute(
        f"""
        INSERT INTO {tgt_schema}.{src_table} ({col_list})
        SELECT {col_list} FROM public.{src_table}
        WHERE lower(table_name) = ANY(%s)
        """,
        (list(yes_names),),
    )
    return cur.rowcount


def copy_full(cur, src_table: str, tgt_schema: str) -> int:
    """Copy all rows from public.src_table."""
    cols = table_columns(cur, src_table)
    col_list = ", ".join(f'"{c}"' for c in cols)
    cur.execute(
        f"""
        INSERT INTO {tgt_schema}.{src_table} ({col_list})
        SELECT {col_list} FROM public.{src_table}
        """
    )
    return cur.rowcount


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--schema", default="adhoc")
    parser.add_argument("--checklist", default="adhoc_table_checklist.xlsx")
    parser.add_argument("--drop", action="store_true", help="Drop and recreate the target schema")
    args = parser.parse_args()

    tgt = args.schema

    if not os.path.exists(args.checklist):
        print(f"ERROR: {args.checklist} not found")
        sys.exit(1)

    print(f"Reading YES tables from: {args.checklist}")
    yes_names = read_yes_tables(args.checklist)
    print(f"  {len(yes_names)} tables marked YES")

    conn = get_conn()
    conn.autocommit = False
    cur = conn.cursor()

    try:
        # Create (or recreate) the target schema
        if args.drop:
            print(f"Dropping schema '{tgt}' if exists...")
            cur.execute(f"DROP SCHEMA IF EXISTS {tgt} CASCADE")

        print(f"Creating schema '{tgt}'...")
        cur.execute(f"CREATE SCHEMA IF NOT EXISTS {tgt}")

        # Discover all schema_* tables in public
        cur.execute(
            """
            SELECT table_name FROM information_schema.tables
            WHERE table_schema = 'public'
              AND table_name LIKE 'schema_%'
              AND table_type = 'BASE TABLE'
            ORDER BY table_name
            """
        )
        all_schema_tables = [r[0] for r in cur.fetchall()]
        print(f"Found schema_* tables in public: {all_schema_tables}")

        for tbl in all_schema_tables:
            # Recreate table structure in target schema
            cur.execute(
                f"CREATE TABLE IF NOT EXISTS {tgt}.{tbl} (LIKE public.{tbl} INCLUDING ALL)"
            )

            if tbl in FILTERED_TABLES:
                print(f"  Copying {tbl} (filtered to YES tables)...")
                n = copy_filtered(cur, tbl, tgt, yes_names)
                print(f"    → {n} rows inserted")
            else:
                print(f"  Copying {tbl} (full copy)...")
                n = copy_full(cur, tbl, tgt)
                print(f"    → {n} rows inserted")

        conn.commit()
        print(f"\nDone. All schema_* tables created under schema '{tgt}'.")

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: {e}")
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
